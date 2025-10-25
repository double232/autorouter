"""
Trial Orders Automation Script - Multi-Provider AI Support
Monitors emails for court documents, downloads PDFs, extracts dates using AI,
and files them in SharePoint with automatic tracking.

Supports: Claude, OpenAI, Gemini, and vLLM (self-hosted)
No Azure AD required - uses IMAP + SharePoint REST API
"""

import os
import re
import base64
import json
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import requests
from bs4 import BeautifulSoup
import win32com.client
from pathlib import Path
from openpyxl import load_workbook, Workbook


class Config:
    """Configuration for the trial orders automation"""

    # Email settings (uses Outlook COM - no credentials needed!)
    EMAIL_SUBJECT_FILTER = "SERVICE OF COURT DOCUMENT"
    EMAIL_FROM_FILTER = "eservice@myflcourtaccess.com"
    EMAIL_FOLDER = "Inbox"  # Or specify a subfolder like "Inbox/Court Documents"

    # Assignment email settings
    ASSIGNMENT_EMAIL_FROM = "sferrara@florida-law.com"
    ASSIGNMENT_SUBJECT_PATTERN = r"Our File no\."  # Pattern to detect assignment emails

    # Local OneDrive paths (defaults - will be overridden from config)
    CASES_FOLDER = r"C:\Users\zucku\OneDrive - Vernis and Bowling\Litigation Operations - Cases"
    TRIAL_TRACKER_EXCEL = r"C:\Users\zucku\OneDrive - Vernis and Bowling\Litigation Operations - Trial Tracker\Trial Tracker.xlsx"

    def __init__(self):
        """Initialize config by reading from config.json and environment variables"""
        # Load from config.json if it exists
        config_data = {}
        config_path = Path("config.json")
        if config_path.exists():
            try:
                with open(config_path, 'r') as f:
                    config_data = json.load(f)
            except Exception as e:
                print(f"WARNING: Could not load config.json: {e}")

        # OneDrive paths (from config or env, fallback to defaults)
        self.CASES_FOLDER = os.getenv("CASES_FOLDER") or config_data.get("cases_folder") or self.CASES_FOLDER
        self.TRIAL_TRACKER_EXCEL = os.getenv("TRIAL_TRACKER_EXCEL") or config_data.get("trial_tracker_excel") or self.TRIAL_TRACKER_EXCEL

        # Test mode (skips PDF validation for testing)
        self.TEST_MODE = os.getenv("TEST_MODE", "").lower() == "true" or config_data.get("test_mode", False)
        self.TEST_PDF_PATH = os.getenv("TEST_PDF_PATH") or config_data.get("test_pdf_path")

        # Defendant to Client mapping (from config)
        self.DEFENDANT_CLIENT_MAP = config_data.get("defendant_client_map", {})
        self.DEFAULT_CLIENT = config_data.get("default_client", "4694")


class EmailClient:
    """Outlook COM email client - uses your existing Outlook installation"""

    def __init__(self, config: Config):
        self.config = config
        self.outlook = None
        self.namespace = None

    def connect(self):
        """Connect to Outlook via COM"""
        try:
            self.outlook = win32com.client.Dispatch("Outlook.Application")
            self.namespace = self.outlook.GetNamespace("MAPI")
            print("Connected to Outlook (using your existing session)")
        except Exception as e:
            raise Exception(
                f"Failed to connect to Outlook: {e}\n\n"
                "Make sure:\n"
                "1. Outlook is installed on this computer\n"
                "2. Outlook is configured with your email account\n"
                "3. You can open Outlook normally\n\n"
                "Note: This uses your existing Outlook installation,\n"
                "so no additional passwords or configuration needed!"
            )

    def disconnect(self):
        """Disconnect from Outlook"""
        # COM objects don't need explicit disconnect
        self.outlook = None
        self.namespace = None

    def get_unread_emails(self) -> List[Dict]:
        """Get unread emails matching the filter criteria"""
        try:
            # Get the Inbox folder
            inbox = self.namespace.GetDefaultFolder(6)  # 6 = olFolderInbox

            emails = []

            # Only search "Daily Mail" subfolder
            folders_to_search = []
            try:
                for subfolder in inbox.Folders:
                    if subfolder.Name.lower() == "daily mail":
                        folders_to_search.append(subfolder)
                        print(f"Found 'Daily Mail' folder")
                        break

                if not folders_to_search:
                    print("WARNING: 'Daily Mail' folder not found in Inbox")
                    return emails
            except Exception as e:
                print(f"Error finding 'Daily Mail' folder: {e}")
                return emails

            print(f"Searching Daily Mail folder for unread emails...")

            # Search each folder
            for folder in folders_to_search:
                try:
                    messages = folder.Items
                    messages.Sort("[ReceivedTime]", True)

                    # Filter for unread emails only (we'll check attachments and subject later)
                    filter_str = "@SQL=\"urn:schemas:httpmail:read\" = 0"
                    filtered_messages = messages.Restrict(filter_str)

                    for msg in filtered_messages:
                        # Process all unread emails in Daily Mail folder
                        try:
                            # Get HTML body (preferred) or plain text
                            body = msg.HTMLBody if hasattr(msg, 'HTMLBody') else msg.Body

                            # Get received time
                            received_time = msg.ReceivedTime if hasattr(msg, 'ReceivedTime') else None

                            print(f"  Found email in folder: {folder.Name}")

                            emails.append({
                                "id": msg.EntryID,
                                "subject": msg.Subject,
                                "body": body,
                                "from": msg.SenderEmailAddress,
                                "received_time": received_time
                            })
                        except Exception as e:
                            print(f"  Warning: Error reading email: {e}")
                            continue
                except:
                    # Skip folders we can't access
                    pass

            return emails

        except Exception as e:
            print(f"Error getting emails: {e}")
            return []

    def mark_as_read(self, email_id: str):
        """Mark email as read"""
        try:
            msg = self.namespace.GetItemFromID(email_id)
            msg.UnRead = False
            msg.Save()
        except Exception as e:
            print(f"  Warning: Error marking email as read: {e}")


class SharePointClient:
    """Local file client - writes to OneDrive folders that sync to SharePoint"""

    def __init__(self, config: Config):
        self.config = config
        self.cases_folder = Path(config.CASES_FOLDER)
        self.trial_tracker_excel = Path(config.TRIAL_TRACKER_EXCEL)

        # Excel and folder caches - loaded on first use
        self._excel_cache_loaded = False
        self._case_index = {}  # {case_number: {client, matter, style}}
        self._party_index = {}  # {party_word: [case_info, ...]}
        self._folder_index = {}  # {client/matter: case_info}

        # Verify folders exist
        if not self.cases_folder.exists():
            raise Exception(f"Cases folder not found: {self.cases_folder}")

        print(f"Using local OneDrive folders")
        print(f"   Cases: {self.cases_folder}")
        print(f"   Tracker: {self.trial_tracker_excel}")

    def _ensure_caches_loaded(self):
        """Load Excel and folder caches on first use"""
        if self._excel_cache_loaded:
            return

        print("Building case index from Excel and folders...")

        # Load Excel index
        try:
            if self.trial_tracker_excel.exists():
                wb = load_workbook(self.trial_tracker_excel, read_only=True, data_only=True)
                if "Lookup Table 2" in wb.sheetnames:
                    ws = wb["Lookup Table 2"]
                    headers = [cell.value for cell in ws[1]]

                    if "Case No." in headers and "Client" in headers and "Matter" in headers:
                        case_no_idx = headers.index("Case No.")
                        client_idx = headers.index("Client")
                        matter_idx = headers.index("Matter")
                        style_idx = headers.index("Style") if "Style" in headers else None

                        for row in ws.iter_rows(min_row=2, values_only=True):
                            case_no = str(row[case_no_idx]).strip() if row[case_no_idx] else None
                            client = str(row[client_idx]) if row[client_idx] else None
                            matter = str(row[matter_idx]) if row[matter_idx] else None
                            style = str(row[style_idx]) if style_idx and row[style_idx] else "Unknown"

                            if case_no and client:
                                self._case_index[case_no] = {
                                    "Client": client,
                                    "Matter": matter or case_no,
                                    "Style": style
                                }
        except Exception as e:
            print(f"  Error loading Excel: {e}")

        # Build folder index and party index
        try:
            for client_folder in self.cases_folder.iterdir():
                if not client_folder.is_dir():
                    continue

                for matter_folder in client_folder.iterdir():
                    if not matter_folder.is_dir():
                        continue

                    client_num = client_folder.name
                    matter_name = matter_folder.name
                    style = matter_name.split(" - ", 1)[1] if " - " in matter_name else "Unknown"

                    folder_key = f"{client_num}/{matter_name}"
                    self._folder_index[folder_key] = {
                        "Client": client_num,
                        "Matter": matter_name,
                        "Style": style,
                        "FolderPath": str(matter_folder),
                        "RelativePath": f"{client_num}\\{matter_name}"
                    }

                    # Index party names
                    if style != "Unknown":
                        for word in style.upper().split():
                            if len(word) >= 3:
                                if word not in self._party_index:
                                    self._party_index[word] = []
                                self._party_index[word].append(self._folder_index[folder_key])

        except Exception as e:
            print(f"  Error building folder index: {e}")

        self._excel_cache_loaded = True
        print(f"  Indexed {len(self._case_index)} Excel entries, {len(self._folder_index)} folders, {len(self._party_index)} party keys")

    def lookup_case_from_excel(self, case_number: str) -> Optional[Dict]:
        """Look up client and matter from court case number in Lookup Table 2 (uses cache)"""
        self._ensure_caches_loaded()

        # Check cache first
        if case_number in self._case_index:
            case_data = self._case_index[case_number]
            print(f"  Found Excel mapping: {case_number} -> Client {case_data['Client']}, Matter {case_data['Matter']}")
            return case_data

        print(f"  WARNING: Case number {case_number} not found in Lookup Table 2")
        return None

    def lookup_case_from_excel_OLD(self, case_number: str) -> Optional[Dict]:
        """OLD: Look up client and matter - replaced by cached version"""
        try:
            if not self.trial_tracker_excel.exists():
                print(f"  WARNING: Trial Tracker Excel not found")
                return None

            wb = load_workbook(self.trial_tracker_excel, read_only=True)

            if "Lookup Table 2" not in wb.sheetnames:
                print(f"  WARNING: 'Lookup Table 2' sheet not found in Excel")
                return None

            ws = wb["Lookup Table 2"]

            # Find header row to get column indices
            headers = [cell.value for cell in ws[1]]

            # Lookup Table 2 has: Attorney, Client, Matter, Style, Claim No., Case No., County, Court, etc.
            if "Case No." not in headers or "Matter" not in headers or "Client" not in headers:
                print(f"  WARNING: Required columns not found in Lookup Table 2")
                return None

            case_no_col = headers.index("Case No.")
            matter_col = headers.index("Matter")
            client_col = headers.index("Client")
            style_col = headers.index("Style") if "Style" in headers else None

            # Search for case number in rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[case_no_col] and str(row[case_no_col]).strip() == case_number.strip():
                    client = str(row[client_col])
                    matter = str(row[matter_col])
                    style = str(row[style_col]) if style_col and row[style_col] else "Unknown"
                    print(f"  Found Excel mapping: {case_number} -> Client {client}, Matter {matter}")
                    return {
                        "Client": client,
                        "Matter": matter,
                        "Style": style
                    }

            print(f"  WARNING: Case number {case_number} not found in Lookup Table 2")
            return None

        except Exception as e:
            print(f"  WARNING: Error reading Excel: {e}")
            return None

    def get_case_by_number(self, case_number: str) -> Optional[Dict]:
        """Look for case folder in local OneDrive - searches using client + matter from Excel"""
        try:
            # First, try to look up the client and matter from the Excel file
            excel_info = self.lookup_case_from_excel(case_number)

            if excel_info:
                # We have client and matter from Excel - search for specific client\matter folder
                client_num = excel_info["Client"]
                matter_num = excel_info["Matter"]

                client_folder = self.cases_folder / client_num
                if not client_folder.exists():
                    print(f"  WARNING: Client folder {client_num} does not exist")
                    return None

                # Look for matter folder starting with the matter number
                for matter_folder in client_folder.iterdir():
                    if matter_folder.is_dir() and matter_folder.name.startswith(matter_num):
                        print(f"  Found case folder: {client_num}\\{matter_folder.name}")

                        matter_name = matter_folder.name  # e.g., "90250040 - Ricciardi"

                        # Extract style from folder name (part after the dash)
                        style = excel_info["Style"]
                        if " - " in matter_name:
                            style = matter_name.split(" - ", 1)[1]

                        return {
                            "Client": client_num,
                            "Matter": matter_name,
                            "Style": style,
                            "CaseNumber": case_number,
                            "FolderPath": str(matter_folder),
                            "RelativePath": f"{client_num}\\{matter_name}"
                        }

                print(f"  WARNING: Matter folder not found in {client_num} for matter {matter_num}")
                return None
            else:
                # No Excel info - fall back to searching all folders
                print(f"  No Excel mapping found, searching all folders for case {case_number}")
                for client_folder in self.cases_folder.iterdir():
                    if not client_folder.is_dir():
                        continue

                    for matter_folder in client_folder.iterdir():
                        if matter_folder.is_dir() and case_number in matter_folder.name:
                            print(f"  Found case folder: {client_folder.name}\\{matter_folder.name}")

                            client_num = client_folder.name
                            matter_name = matter_folder.name

                            style = "Unknown"
                            if " - " in matter_name:
                                style = matter_name.split(" - ", 1)[1]

                            return {
                                "Client": client_num,
                                "Matter": matter_name,
                                "Style": style,
                                "CaseNumber": case_number,
                                "FolderPath": str(matter_folder),
                                "RelativePath": f"{client_num}\\{matter_name}"
                            }

                print(f"  WARNING: No existing folder found for case {case_number}")
                return None

        except Exception as e:
            print(f"Error looking up case: {e}")
            return None

    def upload_file(self, folder_path: str, filename: str, content: bytes) -> str:
        """Save file to local OneDrive folder with duplicate detection"""
        try:
            # folder_path is now a relative Windows path like "272\90250143 - Tomasini\09 Orders"
            # Convert to Path object and combine with cases folder
            local_path = self.cases_folder / folder_path

            # Create folder if it doesn't exist
            local_path.mkdir(parents=True, exist_ok=True)

            # Check for duplicates
            file_path = local_path / filename
            if file_path.exists():
                # File exists - check if identical
                existing_content = file_path.read_bytes()
                if existing_content == content:
                    print(f"  Duplicate detected (identical): {filename} - Skipping")
                    return str(file_path)
                else:
                    print(f"  File exists but content differs: {filename}")
                    print(f"    Existing: {len(existing_content)} bytes, New: {len(content)} bytes")
                    print(f"  Overwriting with new version")

            # Save file
            file_path.write_bytes(content)

            print(f"  Saved to: {file_path}")
            return str(file_path)
        except Exception as e:
            print(f"Error saving file: {e}")
            raise

    def parse_assignment_email(self, subject: str) -> Optional[Dict]:
        """Parse assignment email subject to extract case info
        Format: 'Our File no. 272-90250273 De Leon Reyes, Samuel vs Citizens (001-00-603213) Claim no.:'
        """
        try:
            # Pattern: Our File no. CLIENT-MATTER NAME vs DEFENDANT (CLAIM-NO) Claim no.:
            pattern = r'Our File no\.\s+(\d+)-(\d+)\s+([^(]+?)\s+vs\s+[^(]+\(([^)]+)\)'
            match = re.search(pattern, subject)

            if match:
                client = match.group(1)
                matter = match.group(2)
                style = match.group(3).strip()
                claim_no = match.group(4).strip()

                print(f"  Parsed assignment: Client={client}, Matter={matter}, Style={style}, Claim={claim_no}")

                return {
                    "Client": client,
                    "Matter": matter,
                    "Style": style,
                    "Claim_No": claim_no
                }

            return None

        except Exception as e:
            print(f"  WARNING: Error parsing assignment email: {e}")
            return None

    def create_case_row(self, case_data: Dict):
        """Create new row in Lookup Table 2 from assignment email or court email"""
        try:
            if not self.trial_tracker_excel.exists():
                print(f"  ERROR: Trial Tracker Excel not found")
                return False

            wb = load_workbook(self.trial_tracker_excel)

            if "Lookup Table 2" not in wb.sheetnames:
                print(f"  ERROR: 'Lookup Table 2' sheet not found in Excel")
                return False

            ws = wb["Lookup Table 2"]

            # Get headers
            headers = [cell.value for cell in ws[1]]

            # Build row data with available information
            # Find column indices
            col_indices = {}
            for col_name in ["Attorney", "Client", "Matter", "Style", "Claim No.", "Case No."]:
                if col_name in headers:
                    col_indices[col_name] = headers.index(col_name) + 1

            # Append new row
            new_row = [None] * len(headers)

            # Fill in available data
            if "Attorney" in col_indices:
                new_row[col_indices["Attorney"] - 1] = case_data.get("Attorney", "EAZ")
            if "Client" in col_indices and case_data.get("Client"):
                new_row[col_indices["Client"] - 1] = case_data.get("Client")
            if "Matter" in col_indices and case_data.get("Matter"):
                new_row[col_indices["Matter"] - 1] = case_data.get("Matter")
            if "Style" in col_indices and case_data.get("Style"):
                new_row[col_indices["Style"] - 1] = case_data.get("Style")
            if "Claim No." in col_indices and case_data.get("Claim_No"):
                new_row[col_indices["Claim No."] - 1] = case_data.get("Claim_No")
            if "Case No." in col_indices and case_data.get("Case_No"):
                new_row[col_indices["Case No."] - 1] = case_data.get("Case_No")

            ws.append(new_row)
            wb.save(self.trial_tracker_excel)

            print(f"  Created new row in Lookup Table 2")
            return True

        except Exception as e:
            print(f"  ERROR: Failed to create case row: {e}")
            return False

    def create_trial_order_record(self, data: Dict):
        """Update existing row in 'Lookup Table 2' sheet with trial order information"""
        try:
            if not self.trial_tracker_excel.exists():
                print(f"  ERROR: Trial Tracker Excel not found")
                return

            wb = load_workbook(self.trial_tracker_excel)

            if "Lookup Table 2" not in wb.sheetnames:
                print(f"  ERROR: 'Lookup Table 2' sheet not found in Excel")
                return

            ws = wb["Lookup Table 2"]

            # Find header row to get column indices
            headers = [cell.value for cell in ws[1]]

            # Required columns: Case No., Calendar Call, Trial Date, Order Date
            if "Case No." not in headers:
                print(f"  ERROR: 'Case No.' column not found in Lookup Table 2")
                return

            case_no_col = headers.index("Case No.") + 1  # 1-indexed for openpyxl

            # Optional columns - only update if they exist
            calendar_call_col = headers.index("Calendar Call") + 1 if "Calendar Call" in headers else None
            trial_date_col = headers.index("Trial Date") + 1 if "Trial Date" in headers else None
            order_date_col = headers.index("Order Date") + 1 if "Order Date" in headers else None

            # Find the row for this case number
            case_number = data.get("Case_Number", "")
            row_found = False

            for row_idx in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=case_no_col).value
                if cell_value and str(cell_value).strip() == case_number.strip():
                    row_found = True
                    print(f"  Found existing row for case {case_number} at row {row_idx}")

                    # Update Calendar Call if provided and column exists
                    if calendar_call_col and data.get("Calendar_Call"):
                        ws.cell(row=row_idx, column=calendar_call_col).value = data.get("Calendar_Call")
                        print(f"    Updated Calendar Call: {data.get('Calendar_Call')}")

                    # Update Trial Date if provided and column exists
                    if trial_date_col and data.get("Trial_Start"):
                        ws.cell(row=row_idx, column=trial_date_col).value = data.get("Trial_Start")
                        print(f"    Updated Trial Date: {data.get('Trial_Start')}")

                    # Update Order Date if provided and column exists
                    if order_date_col:
                        # Use current date as Order Date (when the order was processed)
                        order_date = datetime.now().strftime("%Y-%m-%d")
                        ws.cell(row=row_idx, column=order_date_col).value = order_date
                        print(f"    Updated Order Date: {order_date}")

                    break

            if not row_found:
                print(f"  WARNING: No existing row found for case {case_number} in Lookup Table 2")
                print(f"           Please ensure the case is added to the lookup table first")
                return

            # Save workbook
            wb.save(self.trial_tracker_excel)
            print(f"  Updated 'Lookup Table 2' successfully")

        except Exception as e:
            print(f"  ERROR: Error updating Excel: {e}")
            import traceback
            traceback.print_exc()


class PDFProcessor:
    """Process PDFs using regex-based extraction"""

    def __init__(self, config: Config):
        self.config = config

    def _extract_with_regex(self, pdf_content: bytes) -> Dict[str, Optional[str]]:
        """Extract trial dates using regex patterns (fast, free, offline)"""
        try:
            import pdfplumber
            import io
            from datetime import datetime

            # Extract text from PDF
            text = ""
            with pdfplumber.open(io.BytesIO(pdf_content)) as pdf:
                for page in pdf.pages:
                    text += page.extract_text() or ""

            if not text:
                return None  # No text extracted, fall back to AI

            # Normalize text: remove line breaks and "AM"/"PM" within dates
            # e.g., "09-16-\nAM\n2024" -> "09-16-2024"
            text = re.sub(r'(\d{2}-\d{2})-\s*[\n\r]+\s*(?:AM|PM)?\s*[\n\r]*\s*(\d{4})', r'\1-\2', text, flags=re.IGNORECASE)

            # Initialize result
            result = {
                "calendar_call": None,
                "trial_start": None,
                "trial_end": None,
                "document_type": "Other"
            }

            # Detect document type (check UTO first for more specific match)
            if "UNIFORM TRIAL ORDER" in text.upper():
                result["document_type"] = "UTO"
            elif "CASE MANAGEMENT ORDER" in text.upper():
                result["document_type"] = "CMO"

            # Regex patterns for date extraction
            # Pattern 1: Trial Period (UTO) - "TRIAL PERIOD COMMENCING: 11-03-2025 to 11-21-2025"
            trial_period_pattern = r'TRIAL PERIOD COMMENCING:\s*(\d{2}-\d{2}-\d{4})\s*to\s*(\d{2}-\d{2}-\d{4})'
            trial_period_match = re.search(trial_period_pattern, text, re.IGNORECASE)
            if trial_period_match:
                result["trial_start"] = trial_period_match.group(1)
                result["trial_end"] = trial_period_match.group(2)

            # Pattern 2: Calendar Call (UTO) - "CALENDAR CALL: 10-24-2025 at 9:30 AM"
            calendar_call_pattern = r'CALENDAR CALL:\s*(\d{2}-\d{2}-\d{4})'
            calendar_call_match = re.search(calendar_call_pattern, text, re.IGNORECASE)
            if calendar_call_match:
                result["calendar_call"] = calendar_call_match.group(1)

            # Pattern 3: Case Management Conference (CMO) - "on 09-16-2024 9:45" (AM may be stripped)
            # More flexible pattern to handle text/linebreaks between "held"/"on" and date
            # AM/PM is optional because it may have been removed during normalization
            cmc_pattern = r'(?:held|on)[\s\S]{0,200}?(\d{2}-\d{2}-\d{4})[\s\n\r]+\d{1,2}:\d{2}(?:[\s\n\r]*[AP]M)?'
            cmc_match = re.search(cmc_pattern, text, re.IGNORECASE)
            if cmc_match and not result["calendar_call"]:
                result["calendar_call"] = cmc_match.group(1)

            # Convert MM-DD-YYYY to YYYY-MM-DD format for consistency
            def convert_date_format(date_str):
                if not date_str:
                    return None
                try:
                    dt = datetime.strptime(date_str, "%m-%d-%Y")
                    return dt.strftime("%Y-%m-%d")
                except:
                    return date_str  # Return as-is if conversion fails

            result["calendar_call"] = convert_date_format(result["calendar_call"])
            result["trial_start"] = convert_date_format(result["trial_start"])
            result["trial_end"] = convert_date_format(result["trial_end"])

            # Check if we found any dates
            if result["calendar_call"] or result["trial_start"]:
                print(f"  Regex extraction successful: {result['document_type']}")
                return result

            # No dates found with regex
            return None

        except Exception as e:
            print(f"  Regex extraction failed: {e}")
            return None

    def extract_efiling_date(self, pdf_content: bytes) -> Optional[str]:
        """Extract e-filing date from PDF header"""
        try:
            import pdfplumber
            import io
            from datetime import datetime

            # Extract text from first page only (header is typically on page 1)
            with pdfplumber.open(io.BytesIO(pdf_content)) as pdf:
                if not pdf.pages:
                    return None
                text = pdf.pages[0].extract_text() or ""

            # Common e-filing date patterns
            patterns = [
                r'[Ee]-?[Ff]iled?:?\s*(\d{1,2}[-/]\d{1,2}[-/]\d{4})',  # E-Filed: 01/15/2026
                r'[Ff]iled?:?\s*(\d{1,2}[-/]\d{1,2}[-/]\d{4})',        # Filed: 01/15/2026
                r'[Dd]ate [Ff]iled:?\s*(\d{1,2}[-/]\d{1,2}[-/]\d{4})', # Date Filed: 01/15/2026
            ]

            for pattern in patterns:
                match = re.search(pattern, text)
                if match:
                    date_str = match.group(1)
                    # Parse and convert to YYYY.MM.DD format
                    try:
                        # Handle both MM/DD/YYYY and MM-DD-YYYY
                        for fmt in ['%m/%d/%Y', '%m-%d-%Y']:
                            try:
                                dt = datetime.strptime(date_str, fmt)
                                return dt.strftime('%Y.%m.%d')
                            except:
                                continue
                    except:
                        pass

            return None

        except Exception as e:
            print(f"  Warning: Could not extract e-filing date: {e}")
            return None

    def extract_trial_dates(self, pdf_content: bytes) -> Dict[str, Optional[str]]:
        """Extract trial-related dates from PDF using regex"""
        print(f"  Attempting regex extraction...")
        result = self._extract_with_regex(pdf_content)

        if result:
            print(f"  Regex extraction successful")
            return result
        else:
            print(f"  Regex extraction failed - no dates found")
            return {
                "calendar_call": None,
                "trial_start": None,
                "trial_end": None,
                "document_type": "Other"
            }


class TrialOrdersAutomation:
    """Main automation orchestrator"""

    def __init__(self):
        self.config = Config()
        self.email_client = EmailClient(self.config)
        self.sharepoint_client = SharePointClient(self.config)
        self.pdf_processor = PDFProcessor(self.config)

    def extract_caption_info(self, pdf_content: bytes) -> Dict[str, Optional[str]]:
        """Extract case information from pleading caption in PDF"""
        try:
            import pdfplumber
            import io

            # Extract text from first 2 pages (caption is usually on page 1)
            with pdfplumber.open(io.BytesIO(pdf_content)) as pdf:
                if not pdf.pages:
                    return {}

                text = ""
                for page in pdf.pages[:2]:  # First 2 pages
                    text += page.extract_text() or ""

            caption_info = {}

            # Extract case number - common patterns
            case_patterns = [
                r'[Cc]ase [Nn]o\.?:?\s*([A-Z0-9]+)',
                r'[Cc]ase #:?\s*([A-Z0-9]+)',
                r'([0-9]{2}[0-9]{4}[A-Z]{2}[0-9]+[A-Z]+[0-9]*)',  # FL format: 062024CA012345AXXXCE
            ]

            for pattern in case_patterns:
                match = re.search(pattern, text)
                if match:
                    caption_info['case_number'] = match.group(1)
                    break

            # Extract parties (plaintiff vs defendant format)
            # Pattern: "PLAINTIFF NAME\nvs.\nDEFENDANT NAME" or "Plaintiff vs Defendant"
            vs_pattern = r'([A-Z][A-Za-z\s,\.]+?)\s+(?:vs\.?|v\.)\s+([A-Z][A-Za-z\s,\.]+?)(?:\n|,|Case)'
            vs_match = re.search(vs_pattern, text, re.IGNORECASE)
            if vs_match:
                caption_info['plaintiff'] = vs_match.group(1).strip()
                caption_info['defendant'] = vs_match.group(2).strip()

            # Extract court name
            court_pattern = r'IN THE (.+?COURT.+?)(?:IN AND FOR|COUNTY|STATE)'
            court_match = re.search(court_pattern, text, re.IGNORECASE)
            if court_match:
                caption_info['court'] = court_match.group(1).strip()

            return caption_info

        except Exception as e:
            print(f"  Warning: Could not extract caption: {e}")
            return {}

    def find_case_by_party_name(self, party_name: str) -> Optional[Dict]:
        """Search for existing case folder by party name (plaintiff or defendant)"""
        try:
            party_upper = party_name.upper().strip()
            if len(party_upper) < 3:  # Too short to be reliable
                return None

            print(f"  Searching for existing case with party name: {party_name}")

            matches = []

            # Search through all case folders
            for client_folder in self.sharepoint_client.cases_folder.iterdir():
                if not client_folder.is_dir():
                    continue

                for matter_folder in client_folder.iterdir():
                    if not matter_folder.is_dir():
                        continue

                    # Check if party name appears in folder name
                    if party_upper in matter_folder.name.upper():
                        client_num = client_folder.name
                        matter_name = matter_folder.name

                        style = "Unknown"
                        if " - " in matter_name:
                            style = matter_name.split(" - ", 1)[1]

                        matches.append({
                            "Client": client_num,
                            "Matter": matter_name,
                            "Style": style,
                            "CaseNumber": None,
                            "FolderPath": str(matter_folder),
                            "RelativePath": f"{client_num}\\{matter_name}"
                        })

            # Only return if exactly one match found
            if len(matches) == 0:
                print(f"  No matching cases found")
                return None
            elif len(matches) == 1:
                match = matches[0]
                print(f"  Found matching case: {match['RelativePath']}")
                return match
            else:
                print(f"  Found {len(matches)} matching cases - ambiguous, leaving uncategorized")
                return None

        except Exception as e:
            print(f"  Error searching by party name: {e}")
            return None

    def identify_client_from_defendant(self, defendant_text: str) -> str:
        """Identify client number from defendant name using config mappings"""
        defendant_upper = defendant_text.upper()

        # Use defendant mapping from config
        for defendant, client in self.config.DEFENDANT_CLIENT_MAP.items():
            if defendant.upper() in defendant_upper:
                return client

        # Return None if no match (caller will handle with default)
        return None

    def identify_client_from_subject(self, subject: str) -> str:
        """Identify client number from defendant name in email subject"""
        return self.identify_client_from_defendant(subject) or self.config.DEFAULT_CLIENT

    def determine_filing_path(
        self,
        caption_info: Dict,
        case_info: Optional[Dict],
        doc_type: str,
        subfolder: str,
        title: str
    ) -> tuple[str, str]:
        r"""
        Determine filing path based on available information
        Returns: (folder_path, filing_level)

        Filing levels:
        1. "unknown" - No client identified -> Unknown\{title}\
        2. "unsorted_client" - Client only -> {Client}\Unsorted\{title}\
        3. "unsorted_matter" - Client + Matter, no doc type -> {Client}\{Matter}\Unsorted\
        4. "sorted" - Full info -> {Client}\{Matter}\{DocType}\
        """

        # Try to identify client from caption or case_info
        client = None
        if case_info:
            client = case_info.get('Client')

        if not client and caption_info.get('defendant'):
            client = self.identify_client_from_defendant(caption_info['defendant'])

        # Level 1: No client - goes to Unknown
        if not client or client == "Unknown":
            sanitized_title = self.sanitize_filename(title)
            return f"Unknown\\{sanitized_title}", "unknown"

        # We have a client - check if we have matter info
        matter = None
        style = "Unsorted"

        if case_info:
            matter = case_info.get('Matter')
            style = case_info.get('Style', 'Unsorted')

        # Level 2: Client only, no matter - goes to {Client}\Unsorted
        if not matter or matter == "Unknown":
            sanitized_title = self.sanitize_filename(title)
            return f"{client}\\Unsorted\\{sanitized_title}", "unsorted_client"

        # We have client + matter
        relative_path = case_info.get('RelativePath', f"{client}\\{matter}")

        # Level 3: No document type identified - goes to {Client}\{Matter}\Unsorted
        if doc_type == "Unknown" or not subfolder:
            return f"{relative_path}\\Unsorted", "unsorted_matter"

        # Level 4: Full info - goes to {Client}\{Matter}\{DocType}
        return f"{relative_path}\\{subfolder}", "sorted"

    def extract_case_number_from_subject(self, subject: str) -> Optional[str]:
        """Extract case number from email subject
        Handles formats like:
        - CASE NUMBER 062024CA018136AXXXCE
        - CASE NUMBER: 062024CA018136AXXXCE
        - CASE NUMBER:062024CA018136AXXXCE
        """
        # Allow optional colon and flexible whitespace after "CASE NUMBER"
        match = re.search(r'CASE NUMBER\s*:?\s*(\S+)', subject, re.IGNORECASE)
        if match:
            case_num = match.group(1).strip()
            # Remove trailing punctuation
            case_num = re.sub(r'[^\w]+$', '', case_num)
            return case_num
        return None

    def extract_outlook_attachments(self, email_id: str) -> List[Tuple[str, bytes]]:
        """Extract PDF attachments directly from Outlook email, including PDFs inside ZIP files"""
        attachments = []
        try:
            import tempfile
            import os
            import zipfile

            msg = self.email_client.namespace.GetItemFromID(email_id)
            if hasattr(msg, 'Attachments') and msg.Attachments.Count > 0:
                for attachment in msg.Attachments:
                    if not hasattr(attachment, 'FileName'):
                        continue

                    filename = attachment.FileName

                    # Handle PDF attachments
                    if filename.lower().endswith('.pdf'):
                        title = filename[:-4] if filename.lower().endswith('.pdf') else filename

                        # Save to temp file to read content
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
                            tmp_path = tmp.name
                            attachment.SaveAsFile(tmp_path)

                        # Read content
                        with open(tmp_path, 'rb') as f:
                            content = f.read()

                        os.unlink(tmp_path)
                        attachments.append((title, content))
                        print(f"    Found PDF attachment: {filename}")

                    # Handle ZIP attachments - extract PDFs inside
                    elif filename.lower().endswith('.zip'):
                        print(f"    Found ZIP attachment: {filename} - extracting PDFs...")

                        # Save ZIP to temp file
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as tmp:
                            tmp_zip_path = tmp.name
                            attachment.SaveAsFile(tmp_zip_path)

                        # Extract PDFs from ZIP with size limits
                        try:
                            zip_size = os.path.getsize(tmp_zip_path)
                            MAX_ZIP_SIZE = 100 * 1024 * 1024  # 100MB limit
                            MAX_PDF_SIZE = 50 * 1024 * 1024   # 50MB per PDF

                            if zip_size > MAX_ZIP_SIZE:
                                print(f"      WARNING: ZIP too large ({zip_size / 1024 / 1024:.1f}MB > 100MB limit) - skipping")
                            else:
                                with zipfile.ZipFile(tmp_zip_path, 'r') as zip_ref:
                                    for zip_info in zip_ref.namelist():
                                        if zip_info.lower().endswith('.pdf'):
                                            # Check PDF size before extracting
                                            if zip_info.file_size > MAX_PDF_SIZE:
                                                print(f"      Skipping large PDF: {zip_info.filename} ({zip_info.file_size / 1024 / 1024:.1f}MB)")
                                                continue

                                            # Sanitize filename to prevent path traversal
                                            pdf_name = os.path.basename(zip_info.filename)
                                            if not pdf_name or '..' in pdf_name:
                                                print(f"      Skipping suspicious filename: {zip_info.filename}")
                                                continue

                                            pdf_content = zip_ref.read(zip_info)
                                            title = pdf_name[:-4] if pdf_name.lower().endswith('.pdf') else pdf_name
                                            attachments.append((title, pdf_content))
                                            print(f"      Extracted from ZIP: {pdf_name}")
                        except Exception as e:
                            print(f"      Error extracting ZIP: {e}")

                        os.unlink(tmp_zip_path)

        except Exception as e:
            print(f"    Warning: Could not extract attachments: {e}")

        return attachments

    def extract_pdf_links_from_email(self, email_body: str) -> List[Tuple[str, str]]:
        """Extract PDF download links from email HTML body"""
        soup = BeautifulSoup(email_body, 'html.parser')
        links = []

        for link in soup.find_all('a', href=True):
            href = link['href']
            if 'document.nefdd?nai=' in href:
                url = href.replace('&amp;', '&')
                title = link.get_text(strip=True)
                if title.lower().endswith('.pdf'):
                    title = title[:-4]
                links.append((title, url))

        return links

    def detect_document_type(self, pdf_content: bytes, title: str) -> tuple[str, str]:
        """
        Detect document type and return (type, subfolder)
        Returns: (document_type, subfolder_path)
        - Trial Orders: ("UTO"/"CMO", "09 Orders") - processes dates and updates Excel
        - Other documents: (type, subfolder) - just saves to folder
        """
        try:
            import pdfplumber
            import io

            # Extract text from first page for analysis
            with pdfplumber.open(io.BytesIO(pdf_content)) as pdf:
                if not pdf.pages:
                    return ("Unknown", "03 Discovery")
                text = pdf.pages[0].extract_text() or ""

            text_upper = text.upper()
            title_upper = title.upper()

            # Check for Trial Orders (UTO/CMO) - these get special processing
            if "UNIFORM TRIAL ORDER" in text_upper:
                return ("UTO", "09 Orders")
            elif "CASE MANAGEMENT ORDER" in text_upper:
                return ("CMO", "09 Orders")

            # Check for other document types by title/content
            if "PLEADING" in title_upper or "COMPLAINT" in title_upper or "ANSWER" in title_upper:
                return ("Pleading", "02 Pleadings")
            elif "DISCOVERY" in title_upper or "INTERROGATOR" in title_upper or "REQUEST FOR PRODUCTION" in title_upper or "RFP" in title_upper or "RFA" in title_upper:
                return ("Discovery", "03 Discovery")
            elif "DEPOSITION" in title_upper:
                return ("Deposition", "04 Depositions")
            elif "ORDER" in title_upper or "NOTICE OF HEARING" in title_upper:
                return ("Order", "09 Orders")

            # Default to Discovery
            return ("Other", "03 Discovery")

        except Exception as e:
            print(f"  Warning: Could not detect document type: {e}")
            return ("Unknown", "03 Discovery")

    def download_pdf(self, url: str) -> bytes:
        """Download PDF from URL (or use test PDF in test mode)"""
        # Use test PDF if configured
        if self.config.TEST_MODE and self.config.TEST_PDF_PATH:
            print(f"  TEST MODE: Using local PDF: {self.config.TEST_PDF_PATH}")
            from pathlib import Path
            test_path = Path(self.config.TEST_PDF_PATH)
            if test_path.exists():
                content = test_path.read_bytes()
                print(f"  Loaded test PDF ({len(content)} bytes)")
                return content
            else:
                print(f"  WARNING: Test PDF not found, falling back to download")

        # Normal download
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()

        content = response.content

        # Skip validation in test mode
        if self.config.TEST_MODE:
            print(f"  TEST MODE: Skipping PDF validation ({len(content)} bytes)")
            return content

        # Validate it's actually a PDF
        if len(content) < 1000:
            raise Exception(f"Downloaded file too small ({len(content)} bytes) - likely an error page or expired link")

        # Check PDF magic bytes
        if not content.startswith(b'%PDF'):
            raise Exception("Downloaded file is not a PDF - likely an HTML error page or expired link")

        return content

    def sanitize_filename(self, filename: str) -> str:
        """Sanitize filename for SharePoint"""
        invalid_chars = r'[<>:"/\\|?*]'
        filename = re.sub(invalid_chars, '-', filename)
        return filename.strip()

    def process_trial_order(
        self,
        case_info: Dict,
        pdf_content: bytes,
        document_title: str,
        case_number: str,
        document_url: str,
        subfolder: str = "09 Orders"
    ) -> Dict:
        """Process a trial order PDF and create SharePoint records"""

        # Extract dates using AI
        print(f"  Processing PDF with {self.config.AI_PROVIDER.upper()} to extract dates...")
        extracted_data = self.pdf_processor.extract_trial_dates(pdf_content)

        # Extract e-filing date for filename
        efiling_date = self.pdf_processor.extract_efiling_date(pdf_content)
        if efiling_date:
            print(f"  E-Filing date: {efiling_date}")
            date_prefix = efiling_date
        else:
            # Fall back to current date if e-filing date not found
            date_prefix = datetime.now().strftime('%Y.%m.%d')
            print(f"  E-Filing date not found, using current date: {date_prefix}")

        # Build folder path using the relative path from case_info
        client = case_info.get('Client', 'Unknown')
        matter = case_info.get('Matter', 'Unknown')
        style = case_info.get('Style', 'Unknown')

        # Use the RelativePath from case_info (e.g., "272\90250143 - Tomasini")
        relative_path = case_info.get('RelativePath', f"{client}\\{matter}")
        folder_path = f"{relative_path}\\{subfolder}"

        # Create filename with e-filing date prefix
        sanitized_title = self.sanitize_filename(document_title)
        filename = f"{date_prefix} - {sanitized_title}.pdf"

        # Upload to local OneDrive folder
        print(f"  Uploading to: {folder_path}\\{filename}")
        file_url = self.sharepoint_client.upload_file(folder_path, filename, pdf_content)

        # Create tracking record
        record_data = {
            "Title": document_title,
            "Case_Number": case_number,
            "Client": client,
            "Matter": matter,
            "Case_Style": style,
            "Document_Link": file_url,
            "Document_Type": extracted_data.get("document_type", "Other"),
            "Calendar_Call": extracted_data.get("calendar_call"),
            "Trial_Start": extracted_data.get("trial_start"),
            "Trial_End": extracted_data.get("trial_end"),
        }

        # Remove None values
        record_data = {k: v for k, v in record_data.items() if v is not None}

        # Excel updates disabled per user request
        # print(f"  Creating trial order record in SharePoint...")
        # self.sharepoint_client.create_trial_order_record(record_data)

        return {
            "file_url": file_url,
            "extracted_data": extracted_data
        }

    def process_regular_document(
        self,
        case_info: Dict,
        pdf_content: bytes,
        document_title: str,
        case_number: str,
        document_url: str,
        document_type: str,
        subfolder: str
    ) -> Dict:
        """Process a regular (non-trial-order) document - just save to appropriate folder"""

        # Extract e-filing date for filename
        efiling_date = self.pdf_processor.extract_efiling_date(pdf_content)
        if efiling_date:
            print(f"  E-Filing date: {efiling_date}")
            date_prefix = efiling_date
        else:
            # Fall back to current date if e-filing date not found
            date_prefix = datetime.now().strftime('%Y.%m.%d')

        # Build folder path
        client = case_info.get('Client', 'Unknown')
        matter = case_info.get('Matter', 'Unknown')
        relative_path = case_info.get('RelativePath', f"{client}\\{matter}")
        folder_path = f"{relative_path}\\{subfolder}"

        # Create filename with e-filing date prefix
        sanitized_title = self.sanitize_filename(document_title)
        filename = f"{date_prefix} - {sanitized_title}.pdf"

        # Upload to local OneDrive folder
        print(f"  Uploading to: {folder_path}\\{filename}")
        file_url = self.sharepoint_client.upload_file(folder_path, filename, pdf_content)

        # Create tracking record (optional - simpler than trial orders)
        record_data = {
            "Title": document_title,
            "Case_Number": case_number,
            "Client": client,
            "Matter": matter,
            "Document_Link": file_url,
            "Document_Type": document_type,
        }

        # Remove None values
        record_data = {k: v for k, v in record_data.items() if v is not None}

        return {
            "file_url": file_url,
            "document_type": document_type
        }

    def process_email(self, email_data: Dict) -> bool:
        """Process a single email"""
        subject = email_data.get("subject", "")
        body_content = email_data.get("body", "")
        email_id = email_data.get("id", "")
        received_time = email_data.get("received_time")

        print(f"\nProcessing email: {subject}")

        # Check email age - skip if older than 7 days (links expire)
        if received_time:
            # Remove timezone info for comparison (Outlook times are timezone-aware)
            received_naive = received_time.replace(tzinfo=None) if received_time.tzinfo else received_time
            age_days = (datetime.now() - received_naive).days
            print(f"  Email age: {age_days} days")

            if age_days > 7:
                print(f"  WARNING: Email is {age_days} days old - download links likely expired (7 day limit)")
                print(f"  Skipping email...")
                return False

        # Try to extract case number from subject first
        case_number = self.extract_case_number_from_subject(subject)

        if case_number:
            print(f"  Case Number from subject: {case_number}")
        else:
            print("  No case number in subject - will try party name matching")

        # Get case info from SharePoint (if we have a case number)
        case_info = None
        if case_number:
            case_info = self.sharepoint_client.get_case_by_number(case_number)

        # If no case info yet, try matching by party names or claim number in subject
        if not case_info:
            # Try claim number first (format: "2023-632391" or "09678902")
            claim_match = re.search(r'(?:claim|file)[\s#:]+([0-9-]+)', subject, re.IGNORECASE)
            if claim_match:
                claim_num = claim_match.group(1).strip()
                case_info = self.sharepoint_client.get_case_by_number(claim_num)

            # If not found by claim, try party names
            if not case_info:
                # Extract party names (format: "LAST, FIRST v PARTY2")
                # Handle "LAST, FIRST" format to extract just last name
                party_match = re.search(r'([A-Z][A-Za-z]+)(?:,\s*[A-Z][A-Za-z\s]+?)?\s+(?:v\.?s?\.?)\s+([A-Z][A-Za-z\s/]+?)(?:\s+-|\s+/|$)', subject, re.IGNORECASE)
                if party_match:
                    plaintiff_last = party_match.group(1).strip()  # Just the last name
                    defendant = party_match.group(2).strip()

                    # Try to find case by plaintiff last name first
                    case_info = self.find_case_by_party_name(plaintiff_last)

                    # If not found, try defendant
                    if not case_info:
                        case_info = self.find_case_by_party_name(defendant)

        # If still no case info and we have a case number, try to create new row
        if not case_info and case_number:
            print(f"  WARNING: Case {case_number} not found in existing folders")
            print(f"  Attempting to identify client from defendant...")

            # Extract defendant from subject to determine client
            client = self.identify_client_from_subject(subject)
            print(f"  Identified client: {client}")

            # Extract style from subject if possible
            style_match = re.search(r'CASE NUMBER\s+\S+\s+([^,]+)', subject)
            style = style_match.group(1).strip() if style_match else "Unknown"

            print(f"  Creating new row in Lookup Table 2 with available info...")

            # Create minimal case row
            case_data = {
                "Case_No": case_number,
                "Style": style,
                "Attorney": "EAZ",
                "Client": client
            }

            if self.sharepoint_client.create_case_row(case_data):
                print(f"  Row created - continuing with minimal case info")
                # Create minimal case_info for processing with identified client
                case_info = {
                    "Client": client,
                    "Matter": case_number,
                    "Style": style,
                    "CaseNumber": case_number,
                    "FolderPath": None,
                    "RelativePath": f"{client}\\{case_number} - {style}"
                }
            else:
                print(f"  ERROR: Could not create case row")
                return False

        if case_info:
            print(f"  Found case: {case_info.get('Client')} - {case_info.get('Matter')}")
        else:
            print(f"  No case info yet - will extract from PDF captions")

        # Collect PDFs from both sources: download links and direct attachments
        pdfs_to_process = []

        # Source 1: Extract PDF download links from email body (court service emails)
        pdf_links = self.extract_pdf_links_from_email(body_content)
        if pdf_links:
            print(f"  Found {len(pdf_links)} PDF link(s) in email body")
            # Skip first link (ZIP file) and add individual PDFs
            individual_pdfs = pdf_links[1:] if len(pdf_links) > 1 else []
            for title, url in individual_pdfs:
                pdfs_to_process.append(('link', title, url, None))

        # Source 2: Extract direct PDF attachments from Outlook
        print("  Checking for direct PDF attachments...")
        attachments = self.extract_outlook_attachments(email_id)
        if attachments:
            print(f"  Found {len(attachments)} direct attachment(s)")
            for title, content in attachments:
                pdfs_to_process.append(('attachment', title, None, content))

        if not pdfs_to_process:
            print("  WARNING: No PDFs found (no links or attachments)")
            return False

        print(f"  Processing {len(pdfs_to_process)} PDF(s) total")

        # Process each PDF
        documents_processed = 0
        for source_type, title, url, content in pdfs_to_process:
            print(f"\n  Processing: {title}")

            try:
                # Get PDF content based on source
                if source_type == 'link':
                    print(f"  URL: {url}")
                    pdf_content = self.download_pdf(url)
                    print(f"  Downloaded ({len(pdf_content)} bytes)")
                else:  # attachment
                    print(f"  Source: Direct attachment")
                    pdf_content = content
                    print(f"  Size: {len(pdf_content)} bytes")

                # Extract caption info from PDF
                caption_info = self.extract_caption_info(pdf_content)

                # Use caption case number if we don't have one from email
                doc_case_number = case_number
                if not doc_case_number and caption_info.get('case_number'):
                    doc_case_number = caption_info['case_number']
                    print(f"  Case Number from caption: {doc_case_number}")

                    # Try to find case info with caption case number
                    if not case_info:
                        case_info = self.sharepoint_client.get_case_by_number(doc_case_number)

                # Detect document type and routing
                doc_type, subfolder = self.detect_document_type(pdf_content, title)
                print(f"  Document type: {doc_type} -> {subfolder}")

                # Determine filing path using tiered system
                folder_path, filing_level = self.determine_filing_path(
                    caption_info,
                    case_info,
                    doc_type,
                    subfolder,
                    title
                )
                print(f"  Filing level: {filing_level} -> {folder_path}")

                # Extract e-filing date for filename
                efiling_date = self.pdf_processor.extract_efiling_date(pdf_content)
                date_prefix = efiling_date if efiling_date else datetime.now().strftime('%Y.%m.%d')
                if efiling_date:
                    print(f"  E-Filing date: {efiling_date}")

                # Create filename
                sanitized_title = self.sanitize_filename(title)
                filename = f"{date_prefix} - {sanitized_title}.pdf"

                # Upload file
                print(f"  Uploading to: {folder_path}\\{filename}")
                file_url = self.sharepoint_client.upload_file(folder_path, filename, pdf_content)

                # Process trial orders for date extraction (Excel updates disabled)
                if doc_type in ["UTO", "CMO"] and case_info:
                    extracted_data = self.pdf_processor.extract_trial_dates(pdf_content)
                    if any([extracted_data.get('calendar_call'), extracted_data.get('trial_start')]):
                        # Excel updates disabled per user request
                        # record_data = {
                        #     "Title": title,
                        #     "Case_Number": doc_case_number,
                        #     "Document_Type": doc_type,
                        #     "Calendar_Call": extracted_data.get("calendar_call"),
                        #     "Trial_Start": extracted_data.get("trial_start"),
                        #     "Trial_End": extracted_data.get("trial_end"),
                        # }
                        # self.sharepoint_client.create_trial_order_record({k: v for k, v in record_data.items() if v})
                        print(f"  Trial dates extracted:")
                        print(f"     Calendar Call: {extracted_data.get('calendar_call', 'Not found')}")
                        print(f"     Trial Start: {extracted_data.get('trial_start', 'Not found')}")

                print(f"  Filed successfully")
                documents_processed += 1

            except Exception as e:
                print(f"  ERROR: Error processing {title}: {e}")
                import traceback
                traceback.print_exc()
                continue

        if documents_processed == 0:
            print(f"  No documents processed from this email")

        # Mark email as read
        self.email_client.mark_as_read(email_id)

        return True

    def run(self):
        """Main run loop"""
        print("=" * 60)
        print(f"Court Document Automation - Starting")
        print("  - Trial Orders (UTO/CMO): Extract dates + Update Excel")
        print("  - Other Documents: Auto-route to appropriate folders")
        print("=" * 60)

        try:
            # Connect to email
            self.email_client.connect()

            # Get unread emails
            emails = self.email_client.get_unread_emails()
            print(f"\nFound {len(emails)} unread court document emails")

            if not emails:
                print("No emails to process")
                return

            # Process each email
            processed = 0
            for email_data in emails:
                try:
                    if self.process_email(email_data):
                        processed += 1
                except Exception as e:
                    print(f"\nERROR: Error processing email: {e}")
                    continue

            print(f"\n{'=' * 60}")
            print(f"Completed: {processed}/{len(emails)} emails processed successfully")
            print(f"{'=' * 60}")

        except Exception as e:
            print(f"\nFATAL ERROR: {e}")
            raise

        finally:
            # Disconnect from email
            self.email_client.disconnect()


def main():
    """Entry point"""
    automation = TrialOrdersAutomation()
    automation.run()


if __name__ == "__main__":
    main()
