from openpyxl import load_workbook

excel_path = r"C:\Users\zucku\OneDrive - Vernis and Bowling\Litigation Operations - Trial Tracker\Trial Tracker.xlsx"
wb = load_workbook(excel_path, read_only=True)
ws = wb['Lookup Table 2']

headers = [cell.value for cell in ws[1]]
print('Headers:', headers)
print('\nRow 31 (Berhmann):')
for i, cell in enumerate(ws[31]):
    print(f"  {headers[i]}: {cell.value}")

print('\nRow 35 (Ricciardi):')
for i, cell in enumerate(ws[35]):
    print(f"  {headers[i]}: {cell.value}")
